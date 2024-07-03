import csv
import os
import pandas as pd
import pdfplumber

from openpyxl import load_workbook
from pathlib import Path

def pdf_to_excel(pdfile_path: Path, excel_file: Path, overwrite=False):
    """
    Convert tables from a PDF file to an Excel file.

    Args:
        pdfile_path (Path): Path to the input PDF file.
        excel_file (Path): Path to the output Excel file.
        overwrite (bool, optional): Whether to overwrite existing Excel file. Defaults to False.
    
    Returns:
        bool: True if tables were extracted, False otherwise.
    """
    if not overwrite and os.path.exists(excel_file):
        raise ValueError("Excel file already exists. Set overwrite=True to overwrite.")

    # Initialize Excel writer
    writer = pd.ExcelWriter(excel_file, engine='openpyxl')
    tables_extracted = False
    previous_headers = None
    
    # Open PDF file using pdfplumber
    with pdfplumber.open(pdfile_path) as pdf:
        for i, page in enumerate(pdf.pages):
            try:
                # Extract tables from each page
                tables = page.extract_tables()
                if tables:
                    tables_extracted = True
                    for j, table in enumerate(tables):
                        if len(table) > 1:
                            # Handle inconsistent headers by using previous headers
                            if previous_headers and len(table[0]) != len(previous_headers):
                                table[0] = previous_headers
                            else:
                                previous_headers = table[0]
                            
                            # Convert table to DataFrame and write to Excel sheet
                            df = pd.DataFrame(table[1:], columns=table[0])
                            sheet_name = f'Page_{i+1}_Table_{j+1}'
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    # If no tables found, create an empty sheet
                    df = pd.DataFrame({})
                    df.to_excel(writer, sheet_name='No_Tables', index=False)
            except Exception as e:
                print(f"Error extracting table from page {i+1}: {e}")
    
    # Close the Excel writer
    writer.close()
    return tables_extracted

def excel_to_csv(excel_file: Path, csv_file: Path):
    """
    Convert an Excel file to a CSV file.

    Args:
        excel_file (Path): Path to the input Excel file.
        csv_file (Path): Path to the output CSV file.
    """
    try:
        # Load Excel workbook
        wb = load_workbook(excel_file)
        
        # Open CSV file for writing
        with open(csv_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile, delimiter=',')
            
            # Iterate through each sheet in the workbook
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                
                # Iterate through each row in the sheet and write to CSV
                for row in ws.iter_rows(values_only=True):
                    # Replace newline characters with spaces in each cell
                    row = [str(cell).replace('\n', ' ').replace('\r', '') for cell in row]
                    writer.writerow(row)
    
    except Exception as e:
        print(f"Error: {e}")

def csv_to_txt(csv_file: Path, txt_file: Path):
    """
    Convert a CSV file to a text file.

    Args:
        csv_file (Path): Path to the input CSV file.
        txt_file (Path): Path to the output text file.
    """
    try:
        # Open CSV file for reading and text file for writing
        with open(csv_file, 'r', encoding='utf-8') as csvfile, open(txt_file, 'w', encoding='utf-8') as txtfile:
            reader = csv.reader(csvfile, delimiter=',')
            
            # Iterate through each row in the CSV and write to text file
            for row in reader:
                txtfile.write(','.join(row) + '\n')
    
    except Exception as e:
        print(f"Error: {e}")